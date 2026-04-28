import os
import threading
import time
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import httpx
from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, StreamingResponse
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from completar_ubicaciones import BASE_DIR as _SCRIPT_BASE_DIR


DATA_DIR = Path(os.environ.get("DATA_DIR", str(_SCRIPT_BASE_DIR))).resolve()
JOBS_DIR = DATA_DIR / "_jobs"
JOBS_DIR.mkdir(parents=True, exist_ok=True)

ELEMENTOS_SOURCE_URL = os.environ.get("ELEMENTOS_SOURCE_URL", "http://192.168.1.120:8000/download-elementos")

@dataclass
class Job:
    id: str
    status: str  # queued|running|done|error
    error: str | None
    output_path: Path | None
    created_at: float


_jobs: dict[str, Job] = {}
_lock = threading.Lock()


def _job_set(job: Job) -> None:
    with _lock:
        _jobs[job.id] = job


def _job_get(job_id: str) -> Job | None:
    with _lock:
        return _jobs.get(job_id)


def _cleanup_jobs(max_age_s: float = 6 * 60 * 60) -> None:
    now = time.time()
    with _lock:
        to_delete = [jid for jid, j in _jobs.items() if now - j.created_at > max_age_s]
        for jid in to_delete:
            j = _jobs.pop(jid, None)
            if j and j.output_path and j.output_path.exists():
                try:
                    j.output_path.unlink()
                except Exception:
                    pass


def _elementos_path() -> Path:
    return DATA_DIR / "elementos.xlsx"


def _validate_elementos_xlsx(path: Path) -> None:
    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    except (InvalidFileException, ValueError) as e:
        raise HTTPException(status_code=400, detail=f"Archivo Excel inválido: {e}") from e
    sheets = set(wb.sheetnames)
    if "elementos" not in sheets:
        raise HTTPException(status_code=400, detail="El archivo debe tener una hoja llamada 'elementos'")


def _process_job(job_id: str, in_path: Path) -> None:
    job = _job_get(job_id)
    if not job:
        return

    job.status = "running"
    _job_set(job)

    try:
        import completar_ubicaciones as cu

        cu.BASE_DIR = DATA_DIR
        lookup = cu.load_elementos_lookup()
        out = cu.completar_archivo(in_path, lookup=lookup, overwrite=False)

        job.status = "done"
        job.output_path = out
        _job_set(job)
    except Exception as e:
        job.status = "error"
        job.error = str(e)
        _job_set(job)


app = FastAPI()


@app.get("/", response_class=HTMLResponse)
def index() -> Any:
    html_path = Path(__file__).resolve().parent / "static" / "index.html"
    return html_path.read_text(encoding="utf-8")


@app.get("/api/download-elementos")
async def download_elementos():
    """
    Descarga la planilla de elementos desde un servicio externo y la devuelve como attachment.
    Configurable por env var ELEMENTOS_SOURCE_URL.
    """
    try:
        async with httpx.AsyncClient(follow_redirects=True, timeout=60.0) as client:
            r = await client.get(ELEMENTOS_SOURCE_URL)
            r.raise_for_status()
    except httpx.HTTPError as e:
        raise HTTPException(status_code=502, detail=f"No pude descargar elementos: {e}") from e

    media_type = r.headers.get("content-type") or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    headers = {"Content-Disposition": 'attachment; filename="elementos.xlsx"'}
    return StreamingResponse(iter([r.content]), media_type=media_type, headers=headers)


@app.post("/api/upload")
async def upload(file: UploadFile = File(...)) -> dict[str, str]:
    _cleanup_jobs()

    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Subí un archivo .xlsx")

    job_id = uuid.uuid4().hex
    job = Job(id=job_id, status="queued", error=None, output_path=None, created_at=time.time())
    _job_set(job)

    in_path = JOBS_DIR / f"{job_id}_input.xlsx"
    data = await file.read()
    in_path.write_bytes(data)

    t = threading.Thread(target=_process_job, args=(job_id, in_path), daemon=True)
    t.start()

    return {"jobId": job_id}


@app.post("/api/upload-elementos")
async def upload_elementos(file: UploadFile = File(...)) -> dict[str, str]:
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Subí un archivo .xlsx")

    tmp_path = JOBS_DIR / f"{uuid.uuid4().hex}_elementos_tmp.xlsx"
    data = await file.read()
    tmp_path.write_bytes(data)

    _validate_elementos_xlsx(tmp_path)

    target = _elementos_path()
    backup = DATA_DIR / f"elementos_backup_{int(time.time())}.xlsx"

    if target.exists():
        target.replace(backup)
    tmp_path.replace(target)

    return {"ok": "true", "message": "elementos.xlsx actualizado"}


@app.get("/api/status/{job_id}")
def status(job_id: str) -> dict[str, Any]:
    job = _job_get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job inexistente")

    return {
        "jobId": job.id,
        "status": job.status,
        "error": job.error,
        "downloadUrl": f"/api/download/{job.id}" if job.status == "done" else None,
    }


@app.get("/api/download/{job_id}")
def download(job_id: str):
    job = _job_get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job inexistente")
    if job.status != "done" or not job.output_path or not job.output_path.exists():
        raise HTTPException(status_code=409, detail="Todavía no está listo")

    return FileResponse(
        path=str(job.output_path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="ubicaciones_con_ubicacion.xlsx",
    )